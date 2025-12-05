#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Leila Trading Bot Pro - نسخه نهایی
نسخه: 6.0.0
"""

# ========== IMPORTS ==========
import argparse
import asyncio
import functools
import gc
import json
import logging
import os
import re
import smtplib
import sqlite3
import sys
import time
from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Optional, Tuple, Union

import aiohttp
import numpy as np
import pandas as pd
import psutil
import requests
import talib
from aiohttp import ClientTimeout, TCPConnector, web
from cachetools import TTLCache
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from prometheus_client import Counter, Gauge, Histogram, generate_latest

# ========== LOAD ENV ==========
load_dotenv()

# ========== CONFIGURATION ==========
class Config:
    """کلاس کانفیگ یکپارچه و بهبودیافته"""
    
    def __init__(self):
        # مسیرها و دایرکتوری‌ها
        self.OUTPUT_DIR = os.getenv("OUTPUT_DIR", "output")
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)
        self.LOG_FILE = os.path.join(self.OUTPUT_DIR, "bot.log")
        self.DB_PATH = os.path.join(self.OUTPUT_DIR, "signals.db")
        
        # پارامترهای سیگنال و ریسک
        self.MIN_SIGNAL_CONFIDENCE = float(os.getenv("MIN_SIGNAL_CONFIDENCE", "40"))
        self.STRONG_SIGNAL_THRESHOLD = 57
        self.WEAK_SIGNAL_MAX = 50
        self.RISK_PER_TRADE = float(os.getenv("RISK_PER_TRADE", "0.02"))
        self.MAX_POSITION_SIZE = float(os.getenv("MAX_POSITION_SIZE", "0.1"))
        self.INITIAL_BALANCE = float(os.getenv("INITIAL_BALANCE", "10000"))
        
        # تنظیمات اجرا
        self.RUN_INTERVAL = int(os.getenv("RUN_INTERVAL", "3600"))
        self.TIMEFRAMES = tuple(os.getenv("TIMEFRAMES", "15m,30m,1h,4h,1d").split(","))
        self.SYMBOLS_BASE = tuple(os.getenv("SYMBOLS_BASE", "BTC,ETH,SOL,ADA,XRP,DOT,BNB").split(","))
        self.SYMBOLS = [f"{b}/USDT" for b in self.SYMBOLS_BASE]
        
        # امنیت و شبکه
        self.CIRCUIT_BREAKER_THRESHOLD = int(os.getenv("CIRCUIT_BREAKER_THRESHOLD", "5"))
        self.MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))
        self.REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "20"))
        
        # تنظیمات منابع قیمت
        self.PRICE_SOURCE_WEIGHTS = {
            'mexc': 0.20, 'toobit': 0.20, 'coinmarketcap': 0.20, 
            'coingecko': 0.20, 'arzdigital': 0.20,
        }
        
        # تنظیمات منابع خبری
        self.NEWS_SOURCE_WEIGHTS = {
            'newsapi': 0.5,
            'cryptopanic': 0.3,
            'coingecko': 0.2,
        }
        
        # تنظیمات کش
        self.CACHE_TTL = int(os.getenv("CACHE_TTL", "300"))
        self.CACHE_MAXSIZE = int(os.getenv("CACHE_MAXSIZE", "200"))
        
        # تنظیمات ایمیل
        self.EMAIL_ENABLED = os.getenv("EMAIL_ENABLED", "false").lower() == "true"
        self.EMAIL_SMTP_SERVER = os.getenv("EMAIL_SMTP_SERVER", "smtp.gmail.com")
        self.EMAIL_SMTP_PORT = int(os.getenv("EMAIL_SMTP_PORT", "587"))
        self.EMAIL_SENDER = os.getenv("EMAIL_SENDER", "")
        self.EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
        self.EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",")
        
        # تنظیمات پیامک
        self.SMS_ENABLED = os.getenv("SMS_ENABLED", "false").lower() == "true"
        self.SMS_API_KEY = os.getenv("SMS_API_KEY", "")
        self.SMS_RECEIVERS = os.getenv("SMS_RECEIVERS", "").split(",")
        self.SMS_PROVIDER = os.getenv("SMS_PROVIDER", "kavenegar")  # kavenegar, smsir, etc.
        
        # API Keys دیگر
        self.COINGECKO_API_KEY = os.getenv("COINGECKO_API_KEY", "")
        self.NEWSAPI_KEY = os.getenv("NEWSAPI_KEY", "")
        self.CRYPTOPANIC_API_KEY = os.getenv("CRYPTOPANIC_API_KEY", "")
        self.COINMARKETCAP_API_KEY = os.getenv("COINMARKETCAP_API_KEY", "")
        
        # Feature Flags
        self.FEATURE_FLAGS = {
            'advanced_ml': os.getenv("ENABLE_ADVANCED_ML", "true").lower() == "true",
            'news_analysis': os.getenv("ENABLE_NEWS_ANALYSIS", "true").lower() == "true",
            'excel_reports': os.getenv("ENABLE_EXCEL", "true").lower() == "true",
            'health_server': os.getenv("ENABLE_HEALTH_SERVER", "true").lower() == "true",
            'arzdigital_integration': os.getenv("ENABLE_ARZDIGITAL", "true").lower() == "true",
            'email_alerts': os.getenv("ENABLE_EMAIL_ALERTS", "false").lower() == "true",
            'sms_alerts': os.getenv("ENABLE_SMS_ALERTS", "false").lower() == "true",
        }
        
        # تنظیمات SMS
        self.SMS_THRESHOLD = float(os.getenv("SMS_THRESHOLD", "75"))
        self.SMS_MAX_PER_DAY = int(os.getenv("SMS_MAX_PER_DAY", "5"))
        
        # فیلترها
        self.FILTER_CONFIG = {
            'confidence_filter': {
                'min_confidence': self.MIN_SIGNAL_CONFIDENCE,
                'strong_threshold': self.STRONG_SIGNAL_THRESHOLD
            },
            'risk_filter': {
                'max_risk_per_trade': self.RISK_PER_TRADE,
                'max_position_size': self.MAX_POSITION_SIZE
            },
            'volume_filter': {
                'min_volume_btc': 0.1,
                'min_volume_ratio': 0.8
            },
            'timeframe_filter': {
                'required_confirmations': 1,
                'priority_timeframes': ['1h', '4h', '15m', '30m'] 
            }
        }
        
        # ثبت تنظیمات
        self.validate_and_log()
    
    def validate_and_log(self):
        """اعتبارسنجی و ثبت تنظیمات"""
        logger = logging.getLogger("crypto_analyzer")
        
        # بررسی کلیدهای ضروری
        required_keys = {
            "COINMARKETCAP_API_KEY": self.COINMARKETCAP_API_KEY,
            "CRYPTOPANIC_API_KEY": self.CRYPTOPANIC_API_KEY,
            "NEWSAPI_KEY": self.NEWSAPI_KEY,
            "COINGECKO_API_KEY": self.COINGECKO_API_KEY,
        }
        
        for k, v in required_keys.items():
            if not v:
                logger.warning(f"⚠️  کلید محیطی {k} وجود ندارد یا خالی است!")
        
        # بررسی تنظیمات ایمیل
        if self.FEATURE_FLAGS.get('email_alerts'):
            if not all([self.EMAIL_SENDER, self.EMAIL_PASSWORD, self.EMAIL_RECEIVERS]):
                logger.warning("⚠️  قابلیت ایمیل فعال است اما تنظیمات ایمیل کامل نیست!")
        
        # بررسی تنظیمات SMS
        if self.FEATURE_FLAGS.get('sms_alerts'):
            if not all([self.SMS_API_KEY, self.SMS_RECEIVERS]):
                logger.warning("⚠️  قابلیت SMS فعال است اما تنظیمات SMS کامل نیست!")
        
        logger.info(f"✅ تنظیمات بارگذاری شد. {len(self.SYMBOLS)} نماد فعال")
        logger.info(f"📁 پوشه خروجی: {self.OUTPUT_DIR}")
        logger.info(f"📊 تایم‌فریم‌ها: {', '.join(self.TIMEFRAMES)}")
        logger.info(f"📧 قابلیت ایمیل: {'فعال' if self.EMAIL_ENABLED else 'غیرفعال'}")
        logger.info(f"📱 قابلیت SMS: {'فعال' if self.SMS_ENABLED else 'غیرفعال'}")

config = Config()

# ========== LOGGING ==========
logger = logging.getLogger("crypto_analyzer")
logger.setLevel(logging.INFO)
logger.propagate = False

fmt = logging.Formatter(
    "%(asctime)s | %(levelname)-8s | %(message)s",
    "%Y-%m-%d %H:%M:%S"
)

# File Handler
fh = RotatingFileHandler(
    config.LOG_FILE,
    maxBytes=5_000_000,
    backupCount=3,
    encoding="utf-8"
)
fh.setFormatter(fmt)

# Console Handler
ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(fmt)

logger.handlers.clear()
logger.addHandler(fh)
logger.addHandler(ch)

logger.info("🚀 Leila Trading Bot Pro (نسخه 6.0) شروع به کار کرد")

# ========== PROMETHEUS METRICS ==========
try:
    REQUESTS_TOTAL = Counter("requests_total", "Total HTTP requests", ["method", "endpoint", "status"])
    REQUEST_DURATION = Histogram("request_duration_seconds", "HTTP request duration seconds")
    PRICE_SOURCE_SUCCESS = Gauge("price_source_success_rate", "Success rate per price source", ["source"])
    OHLCV_SOURCE_SUCCESS = Gauge("ohlcv_source_success", "OHLCV source success flag", ["source"])
    OHLCV_FETCH_FAILURES = Counter("ohlcv_fetch_failures_total", "OHLCV fetch failures", ["symbol"])
    ACTIVE_SIGNALS = Gauge("active_signals", "Number of active signals")
    CACHE_HIT_RATE = Gauge("cache_hit_rate", "Cache hit rate")
    EMAILS_SENT = Counter("emails_sent_total", "Total emails sent")
    SMS_SENT = Counter("sms_sent_total", "Total SMS sent")
    SIGNAL_QUALITY = Gauge("signal_quality", "Average signal confidence")
except Exception as e:
    logger.debug(f"خطای اولیه‌سازی متریک‌ها: {e}")
    REQUESTS_TOTAL = None
    REQUEST_DURATION = None
    PRICE_SOURCE_SUCCESS = None
    OHLCV_SOURCE_SUCCESS = None
    OHLCV_FETCH_FAILURES = None
    ACTIVE_SIGNALS = None
    CACHE_HIT_RATE = None
    EMAILS_SENT = None
    SMS_SENT = None
    SIGNAL_QUALITY = None

# ========== UTILITY FUNCTIONS ==========
class Utils:
    """توابع کمکی عمومی"""
    
    @staticmethod
    def safe_float(value, default=0.0):
        """تبدیل امن به float"""
        try:
            return float(value) if value is not None else default
        except (ValueError, TypeError):
            return default
    
    @staticmethod
    def safe_get(series, index=-1, default=0):
        """دریافت امن از سری‌ها"""
        try:
            if series is None:
                return default
            if isinstance(series, (list, tuple, np.ndarray)):
                if len(series) == 0:
                    return default
                return series[index]
            if hasattr(series, "empty") and series.empty:
                return default
            idx = len(series) + index if index < 0 else index
            if idx < 0 or idx >= len(series):
                return default
            value = series.iloc[idx] if hasattr(series, "iloc") else series[idx]
            return value if not pd.isna(value) else default
        except Exception:
            return default
    
    @staticmethod
    def fmt_num(value, digits=6, default="-"):
        """قالب‌بندی اعداد"""
        try:
            v = float(value)
            return f"{v:.{digits}f}"
        except Exception:
            return default
    
    @staticmethod
    def validate_symbol(symbol: str) -> bool:
        """اعتبارسنجی فرمت سیمبل"""
        return re.match(r"^[A-Z]+/[A-Z]+$", symbol) is not None
    
    @staticmethod
    def calculate_atr(df: pd.DataFrame, period: int = 14) -> float:
        """محاسبه ATR"""
        try:
            atr = talib.ATR(df['high'], df['low'], df['close'], timeperiod=period)
            return Utils.safe_get(atr, -1, 0)
        except Exception:
            return (df['high'] - df['low']).tail(period).mean() if not df.empty else 0

# ========== MARKET STATE DETECTOR ==========
class MarketStateDetector:
    """
    تشخیص حالت بازار: TREND یا RANGE
    ترکیب ADX + ATR + EMA کراس‌ها
    """
    def __init__(self, adx_period: int = 14, atr_period: int = 14, ema_fast: int = 9, ema_slow: int = 21):
        self.adx_period = adx_period
        self.atr_period = atr_period
        self.ema_fast = ema_fast
        self.ema_slow = ema_slow

    def detect(self, df: pd.DataFrame) -> Dict[str, Any]:
        # محاسبه اندیکاتورها و برگرداندن وضعیت بازار
        ...

detector = MarketStateDetector()

#-----------------MarketStateDetector---------------------------------------------------------------------
def analyze_signal(df: pd.DataFrame, symbol: str, timeframe: str) -> Optional[Dict[str, Any]]:
    market_info = detector.detect(df)

    signal = {
        "symbol": symbol,
        "timeframe": timeframe,
        "signal": "BUY" if market_info["state"] == "TREND" else "HOLD",
        "confidence": Utils.safe_float(market_info["adx"], 0.0),
        "entry_price": Utils.safe_float(Utils.safe_get(df['close'], -1, 0), 0.0),
        "stop_loss": Utils.safe_float(Utils.safe_get(df['low'], -1, 0), 0.0),
        "take_profit": Utils.safe_float(Utils.safe_get(df['high'], -1, 0), 0.0),
        "market_state": market_info["state"],
        "trend_score": market_info["trend_score"]
    }

    # 🆕 فیلتر اعتماد
    if signal["confidence"] < 40.0:
        logger.info(f"⏩ سیگنال {symbol} ({timeframe}) حذف شد: اعتماد {signal['confidence']:.1f}% کمتر از آستانه")
        return None

    # ذخیره در دیتابیس
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO signals 
            (ts, symbol, timeframe, signal, confidence, price, sl, tp, 
             market_state, trend_score)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.now().isoformat(),
            signal["symbol"],
            signal["timeframe"],
            signal["signal"],
            signal["confidence"],
            signal["entry_price"],
            signal["stop_loss"],
            signal["take_profit"],
            signal["market_state"],
            signal["trend_score"]
        ))

    return signal

# ========== SMART CACHE ==========
class SmartCache:
    """کش هوشمند با ردیابی نرخ موفقیت"""
    def __init__(self, maxsize=200, ttl=300):
        self.cache = TTLCache(maxsize=maxsize, ttl=ttl)
        self.success_count = defaultdict(int)
        self.total_requests = defaultdict(int)
        self.hits = 0
        self.misses = 0
        self._lock = asyncio.Lock()

    def get_success_rate(self, key):
        if self.total_requests[key] == 0:
            return 0.0
        return self.success_count[key] / self.total_requests[key]

    def record_success(self, key):
        self.success_count[key] += 1
        self.total_requests[key] += 1

    def record_failure(self, key):
        self.total_requests[key] += 1

    def get_hit_rate(self):
        total = self.hits + self.misses
        return self.hits / total if total > 0 else 0.0

    async def get_or_set(self, key, coroutine, *args, **kwargs):
        async with self._lock:
            cached_value = self.get(key)
            if cached_value is not None:
                return cached_value
            result = await coroutine(*args, **kwargs)
            if result is not None:
                self[key] = result
            return result

    def __contains__(self, key):
        return key in self.cache

    def __getitem__(self, key):
        if key in self.cache:
            self.hits += 1
            return self.cache[key]
        self.misses += 1
        raise KeyError(key)

    def __setitem__(self, key, value):
        self.cache[key] = value

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default

# ========== CIRCUIT BREAKER ==========
class CircuitBreaker:
    """مدار قطع کننده برای مدیریت خطاهای متوالی"""
    def __init__(self, failure_threshold=5, reset_timeout=60):
        self.failure_threshold = failure_threshold
        self.reset_timeout = reset_timeout
        self.failure_count = 0
        self.last_failure_time = None
        self.state = "CLOSED"

    async def call(self, coro):
        if self.state == "OPEN":
            if self.last_failure_time and (time.time() - self.last_failure_time > self.reset_timeout):
                self.state = "HALF_OPEN"
                logger.debug("مدار به HALF_OPEN رفت")
            else:
                raise Exception("Circuit breaker باز است - سرویس موقتاً در دسترس نیست")
        try:
            result = await coro
            self._on_success()
            return result
        except Exception:
            self._on_failure()
            raise

    def _on_success(self):
        self.failure_count = 0
        self.last_failure_time = None
        self.state = "CLOSED"

    def _on_failure(self):
        self.failure_count += 1
        self.last_failure_time = time.time()
        if self.failure_count >= self.failure_threshold:
            self.state = "OPEN"
            logger.warning(f"مدار حفاظتی پس از {self.failure_count} خطا باز شد")

# ========== DATABASE MANAGER ==========
@contextmanager
def get_db_connection():
    conn = sqlite3.connect(config.DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def init_db():
    """ایجاد جداول دیتابیس"""
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS signals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT,
                symbol TEXT NOT NULL,
                timeframe TEXT NOT NULL,
                signal TEXT,
                score REAL,
                confidence REAL,
                price REAL,
                sl REAL,
                tp REAL,
                news_score REAL,
                price_rel REAL,
                news_rel REAL,
                ml_agreement INTEGER,
                ml_confidence REAL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                arz_price REAL,
                price_diff_percent REAL,
                entry_adjusted REAL,
                rule_name TEXT,
                rule_side TEXT,
                rule_entry REAL,
                rule_sl REAL,
                rule_tp REAL,
                rule_confidence REAL,
                           market_state TEXT,             
                     trend_score REAL,              
                UNIQUE(symbol, timeframe, ts)
            )
        """)
        
        # ایجاد جدول لاگ عملکرد
        cur.execute("""
            CREATE TABLE IF NOT EXISTS performance_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                total_symbols INTEGER,
                total_signals INTEGER,
                avg_confidence REAL,
                execution_time REAL,
                memory_usage_mb REAL
            )
        """)
        
        # ایجاد جدول لاگ ارسال‌ها
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notification_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                platform TEXT,
                symbol TEXT,
                confidence REAL,
                message TEXT,
                success BOOLEAN
            )
        """)
        
        logger.info("✅ دیتابیس آماده شد")
def migrate_db_signals():
    """افزودن ستون‌های جدید به جدول signals در صورت نبودن"""
    missing_cols = [
        ("rule_name", "TEXT"),
        ("rule_side", "TEXT"),
        ("rule_entry", "REAL"),
        ("rule_sl", "REAL"),
        ("rule_tp", "REAL"),
        ("rule_confidence", "REAL"),
        ("market_state", "TEXT"),
        ("trend_score", "REAL"),
        ("arz_price", "REAL"),
        ("price_diff_percent", "REAL"),
        ("entry_adjusted", "REAL"),
    ]

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(signals)")
        existing = {row["name"] for row in cur.fetchall()}
        for col, coltype in missing_cols:
            if col not in existing:
                cur.execute(f"ALTER TABLE signals ADD COLUMN {col} {coltype}")
                logger.info(f"✅ ستون جدید اضافه شد: {col} ({coltype})")


# ========== HTTP CLIENT ==========
USER_AGENT = "LeilaTraderPro/6.0"
HEADERS_DEFAULT = {"User-Agent": USER_AGENT, "Accept": "application/json"}
http_cb = CircuitBreaker(failure_threshold=config.CIRCUIT_BREAKER_THRESHOLD, reset_timeout=60)

async def http_get(
    session: aiohttp.ClientSession,
    url: str,
    params: Optional[Dict] = None,
    headers: Optional[Dict] = None,
    retries: int = config.MAX_RETRIES,
    timeout: int = config.REQUEST_TIMEOUT,
) -> Optional[Dict]:
    """درخواست HTTP با قابلیت بازخوانی"""
    async def _do():
        merged_headers = {**HEADERS_DEFAULT, **(headers or {})}
        for i in range(retries):
            start_time = time.time()
            try:
                async with session.get(
                    url,
                    params=params,
                    headers=merged_headers,
                    timeout=ClientTimeout(total=timeout),
                ) as resp:
                    text = await resp.text()
                    
                    # ثبت متریک
                    if REQUESTS_TOTAL:
                        REQUESTS_TOTAL.labels(method="GET", endpoint=url, status=resp.status).inc()
                    
                    if resp.status == 200:
                        if REQUEST_DURATION:
                            REQUEST_DURATION.observe(time.time() - start_time)
                        try:
                            return await resp.json()
                        except Exception:
                            logger.debug(f"خطای JSON برای {url}: {text[:200]}")
                            return None
                    else:
                        logger.debug(f"پاسخ {resp.status} از {url}: {text[:200]}")
                        
            except Exception as e:
                logger.debug(f"خطای GET برای {url}: {e}")
                
            if i < retries - 1:
                await asyncio.sleep(0.9 * (2**i))
        return None
    
    return await http_cb.call(_do())

# ========== SMS MANAGER ==========
class SMSManager:
    """مدیریت ارسال پیامک"""
    
    def __init__(self, config):
        self.config = config
        self.today_sms_count = 0
        self.last_sms_date = None
    
    def reset_daily_counter(self):
        """بازنشانی شمارنده روزانه"""
        today = datetime.now().date()
        if self.last_sms_date != today:
            self.today_sms_count = 0
            self.last_sms_date = today
    
    def send_sms_kavenegar(self, receptor: str, message: str) -> bool:
        """ارسال پیامک از طریق کاوه‌نگار"""
        try:
            url = f"https://api.kavenegar.com/v1/{self.config.SMS_API_KEY}/sms/send.json"
            payload = {
                "receptor": receptor,
                "message": message
            }
            resp = requests.post(url, data=payload, timeout=10)
            
            if resp.status_code == 200:
                logger.info(f"✅ پیامک به {receptor} ارسال شد")
                return True
            else:
                logger.warning(f"⚠️  خطا در ارسال پیامک: {resp.status_code}")
                return False
                
        except Exception as e:
            logger.error(f"❌ خطای ارسال پیامک: {e}")
            return False
    
    def send_sms(self, message: str) -> Dict[str, List[str]]:
        """ارسال پیامک به همه گیرندگان"""
        self.reset_daily_counter()
        
        if not self.config.SMS_ENABLED:
            logger.debug("قابلیت SMS غیرفعال است")
            return {"success": [], "failed": []}
        
        if self.today_sms_count >= self.config.SMS_MAX_PER_DAY:
            logger.warning(f"⚠️  حد مجاز پیامک روزانه ({self.config.SMS_MAX_PER_DAY}) رسیده است")
            return {"success": [], "failed": self.config.SMS_RECEIVERS}
        
        success = []
        failed = []
        
        for receptor in self.config.SMS_RECEIVERS:
            if self.today_sms_count >= self.config.SMS_MAX_PER_DAY:
                failed.append(receptor)
                continue
                
            if self.config.SMS_PROVIDER == "kavenegar":
                result = self.send_sms_kavenegar(receptor, message)
            else:
                logger.warning(f"⚠️  ارائه‌دهنده SMS ناشناخته: {self.config.SMS_PROVIDER}")
                result = False
            
            if result:
                success.append(receptor)
                self.today_sms_count += 1
                
                # ثبت متریک
                if SMS_SENT:
                    SMS_SENT.inc()
            else:
                failed.append(receptor)
        
        logger.info(f"📱 ارسال پیامک: {len(success)} موفق، {len(failed)} ناموفق")
        return {"success": success, "failed": failed}
    
    def format_signal_sms(self, signal: Dict) -> str:
        """قالب‌بندی پیامک سیگنال"""
        symbol = signal.get('symbol', '')
        signal_type = signal.get('signal', '')
        confidence = signal.get('confidence', 0)
        entry = signal.get('entry_price', 0)
        
        # کوتاه کردن برای SMS
        if "BUY" in signal_type:
            action = "خرید"
            emoji = "🟢"
        elif "SELL" in signal_type:
            action = "فروش"
            emoji = "🔴"
        else:
            action = "منتظر"
            emoji = "⚪"
        
        message = f"{emoji} سیگنال {action}\n"
        message += f"نماد: {symbol}\n"
        message += f"اعتماد: {confidence:.0f}%\n"
        message += f"ورود: {Utils.fmt_num(entry)}\n"
        message += f"زمان: {datetime.now().strftime('%H:%M')}"
        
        return message

sms_manager = SMSManager(config)

# ========== EMAIL MANAGER ==========
class EmailManager:
    """مدیریت ارسال ایمیل"""

    def __init__(self, config):
        self.config = config

    def send_email(self, subject: str, body: str, html_body: str = None) -> bool:
        """ارسال ایمیل"""
        if not self.config.EMAIL_ENABLED:
            logger.debug("قابلیت ایمیل غیرفعال است")
            return False

        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = self.config.EMAIL_SENDER
            msg['To'] = ", ".join(self.config.EMAIL_RECEIVERS)

            # متن ساده
            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            # متن HTML (اختیاری)
            if html_body:
                msg.attach(MIMEText(html_body, 'html', 'utf-8'))

            with smtplib.SMTP(self.config.EMAIL_SMTP_SERVER, self.config.EMAIL_SMTP_PORT) as server:
                server.starttls()
                server.login(self.config.EMAIL_SENDER, self.config.EMAIL_PASSWORD)
                server.send_message(msg)

            logger.info("✅ ایمیل ارسال شد")

            if EMAILS_SENT:
                EMAILS_SENT.inc()

            return True

        except Exception as e:
            logger.error(f"❌ خطای ارسال ایمیل: {e}")
            return False

    def format_signal_email(self, signal: Dict) -> Tuple[str, str, str]:
        """قالب‌بندی ایمیل سیگنال"""
        symbol = signal.get('symbol', '')
        timeframe = signal.get('timeframe', '')
        signal_type = signal.get('signal', '')
        confidence = signal.get('confidence', 0)
        entry = signal.get('entry_price', 0)
        sl = signal.get('stop_loss', 0)
        tp = signal.get('take_profit', 0)
        arz_diff = signal.get('price_diff_percent', 0)
        market_state = signal.get('market_state', 'UNKNOWN')
        trend_score = signal.get('trend_score', 0)

        subject = f"🚀 سیگنال {signal_type} - {symbol} ({timeframe})"

        # متن ساده
        body = f"""
سیگنال جدید شناسایی شد:

📊 نماد: {symbol}
⏰ تایم‌فریم: {timeframe}
🚦 سیگنال: {signal_type}
🎯 اعتماد: {confidence:.1f}%
📈 وضعیت بازار: {market_state} (امتیاز: {trend_score})

💰 ورود: {Utils.fmt_num(entry)}
📉 حد ضرر: {Utils.fmt_num(sl)}
📈 حد سود: {Utils.fmt_num(tp)}

🕒 زمان: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        if arz_diff:
            body += f"\n🔁 تفاوت ArzDigital: {arz_diff:.2f}%"

        # HTML
        html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; }}
        .signal {{ padding: 20px; border-radius: 10px; }}
        .buy {{ background-color: #d4edda; }}
        .sell {{ background-color: #f8d7da; }}
        .hold {{ background-color: #fff3cd; }}
    </style>
</head>
<body>
    <div class="signal {'buy' if 'BUY' in signal_type else 'sell' if 'SELL' in signal_type else 'hold'}">
        <h2>🚀 سیگنال جدید شناسایی شد</h2>
        <p><strong>📊 نماد:</strong> {symbol}</p>
        <p><strong>⏰ تایم‌فریم:</strong> {timeframe}</p>
        <p><strong>🚦 سیگنال:</strong> {signal_type}</p>
        <p><strong>🎯 اعتماد:</strong> {confidence:.1f}%</p>
        <p><strong>📈 وضعیت بازار:</strong> {market_state} (امتیاز: {trend_score})</p>
        <hr>
        <p><strong>💰 ورود:</strong> {Utils.fmt_num(entry)}</p>
        <p><strong>📉 حد ضرر:</strong> {Utils.fmt_num(sl)}</p>
        <p><strong>📈 حد سود:</strong> {Utils.fmt_num(tp)}</p>
        <hr>
        <p><strong>🕒 زمان:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        {f'<p><strong>🔁 تفاوت ArzDigital:</strong> {arz_diff:.2f}%</p>' if arz_diff else ''}
    </div>
</body>
</html>
"""
        return subject, body, html_body

# ========== PRICE FETCHERS ==========
PRICE_CACHE = SmartCache(maxsize=300, ttl=300)

def async_cached(cache: Union[SmartCache, TTLCache]):
    """دکوریتور برای کش کردن توابع async"""
    def decorator(func):
        @functools.wraps(func)
        async def wrapper(*args, **kwargs):
            try:
                safe_args = [a for a in args if not isinstance(a, aiohttp.ClientSession)]
                key = json.dumps({"fn": func.__name__, "args": safe_args, "kwargs": kwargs}, default=str, sort_keys=True)
            except Exception:
                key = func.__name__ + str(args) + str(kwargs)
            
            try:
                if key in cache:
                    return cache[key]
            except Exception:
                pass
            
            result = await func(*args, **kwargs)
            try:
                cache[key] = result
            except Exception:
                pass
            return result
        return wrapper
    return decorator

@async_cached(PRICE_CACHE)
async def fetch_price_arzdigital(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:

    """دریافت قیمت از ArzDigital.com"""
    try:
        symbol_map = {
            "BTC/USDT": "bitcoin",
            "ETH/USDT": "ethereum",
            "BNB/USDT": "binance-coin",
            "ADA/USDT": "cardano",
            "SOL/USDT": "solana",
            "XRP/USDT": "ripple",
            "DOT/USDT": "polkadot"
            
        }
        
        coin_slug = symbol_map.get(symbol)
        if not coin_slug:
            return None
        
        url = f"https://api.arzdigital.com/coins/{coin_slug}/"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json",
            "Referer": "https://arzdigital.com/",
        }
        
        data = await http_get(session, url, headers=headers, timeout=15)
        
        if data:
            if "current_price" in data:
                price = float(data["current_price"])
                PRICE_CACHE.record_success("arzdigital")
                return price
            elif "price" in data:
                price = float(data["price"])
                PRICE_CACHE.record_success("arzdigital")
                return price
                
        PRICE_CACHE.record_failure("arzdigital")
        return None
                
    except Exception as e:
        logger.debug(f"خطای ArzDigital برای {symbol}: {e}")
        PRICE_CACHE.record_failure("arzdigital")
        return None

@async_cached(PRICE_CACHE)
async def fetch_price_mexc(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    """دریافت قیمت از MEXC"""
    try:
        market = symbol.replace("/", "").upper()
        url = f"https://api.mexc.com/api/v3/ticker/price?symbol={market}"
        data = await http_get(session, url, timeout=10)
        if data and "price" in data:
            PRICE_CACHE.record_success("mexc")
            return float(data["price"])
    except Exception as e:
        logger.debug(f"خطای MEXC برای {symbol}: {e}")
    PRICE_CACHE.record_failure("mexc")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_toobit(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    """دریافت قیمت از Toobit"""
    try:
        market = symbol.replace("/", "")
        url = f"https://api.toobit.com/v5/market/tickers?category=spot&symbol={market}"
        data = await http_get(session, url, timeout=10)
        if (data and isinstance(data, dict) and "result" in data and 
            isinstance(data["result"], dict) and "list" in data["result"] and 
            isinstance(data["result"]["list"], list) and len(data["result"]["list"]) > 0 and 
            "lastPrice" in data["result"]["list"][0]):
            PRICE_CACHE.record_success("toobit")
            return float(data["result"]["list"][0]["lastPrice"])
    except Exception as e:
        logger.debug(f"خطای Toobit برای {symbol}: {e}")
    PRICE_CACHE.record_failure("toobit")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_coingecko(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    """دریافت قیمت از CoinGecko"""
    try:
        coin_map = {
            'BTC/USDT': 'bitcoin', 'ETH/USDT': 'ethereum', 'BNB/USDT': 'binancecoin',
            'ADA/USDT': 'cardano', 'SOL/USDT': 'solana', 'XRP/USDT': 'ripple',
            'DOT/USDT': 'polkadot'            
        }
        coin_id = coin_map.get(symbol)
        if not coin_id:
            return None
        
        url = "https://api.coingecko.com/api/v3/simple/price"
        params = {'ids': coin_id, 'vs_currencies': 'usd'}
        headers = {"x-cg-demo-api-key": config.COINGECKO_API_KEY} if config.COINGECKO_API_KEY else {}
        
        data = await http_get(session, url, params=params, headers=headers, timeout=10)
        if data and coin_id in data:
            PRICE_CACHE.record_success("coingecko")
            return float(data[coin_id]['usd'])
    except Exception as e:
        logger.debug(f"خطای CoinGecko برای {symbol}: {e}")
    PRICE_CACHE.record_failure("coingecko")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_coinmarketcap(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    """دریافت قیمت از CoinMarketCap"""
    try:
        base = symbol.split('/')[0]
        url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
        headers = {"X-CMC_PRO_API_KEY": config.COINMARKETCAP_API_KEY} if config.COINMARKETCAP_API_KEY else {}
        params = {"symbol": base, "convert": "USD"}
        
        data = await http_get(session, url, params=params, headers=headers, timeout=10)
        if data and 'data' in data and base in data['data']:
            PRICE_CACHE.record_success("coinmarketcap")
            return float(data['data'][base]['quote']['USD']['price'])
    except Exception as e:
        logger.debug(f"خطای CMC برای {symbol}: {e}")
    PRICE_CACHE.record_failure("coinmarketcap")
    return None

async def fetch_price_weighted(session: aiohttp.ClientSession, symbol: str):
    """دریافت قیمت وزندهی شده از تمام منابع"""
    try:
        tasks = {
            'mexc': fetch_price_mexc(session, symbol),
            'toobit': fetch_price_toobit(session, symbol),
            'coingecko': fetch_price_coingecko(session, symbol),
            'coinmarketcap': fetch_price_coinmarketcap(session, symbol),
            'arzdigital': fetch_price_arzdigital(session, symbol),
        }
        
        results = await asyncio.gather(*tasks.values(), return_exceptions=True)
        
        active_sources = {}
        for (name, _), result in zip(tasks.items(), results):
            if isinstance(result, Exception) or result is None:
                continue
            price = result
            base_w = config.PRICE_SOURCE_WEIGHTS.get(name, 0.15)
            sr = PRICE_CACHE.get_success_rate(name)
            dyn_w = max(0.1, min(0.4, base_w * (0.8 + 0.4 * sr)))
            active_sources[name] = (float(price), dyn_w)

        if not active_sources:
            logger.warning(f"همه منابع قیمت برای {symbol} ناموفق بودند")
            
            # Fallback: سعی مجدد در منابع اصلی
            retry_cg = await fetch_price_coingecko(session, symbol)
            retry_cmc = await fetch_price_coinmarketcap(session, symbol)
            candidates = [p for p in [retry_cg, retry_cmc] if isinstance(p, (int, float))]
            
            if candidates:
                final_price = float(np.mean(candidates))
                reliability = 0.25
                logger.info(f"Fallback قیمت برای {symbol}: {Utils.fmt_num(final_price)}")
                return final_price, reliability, {}, None
            else:
                return None, 0.0, {}, None

        # اگر تنها یک منبع فعال باشد
        if len(active_sources) == 1:
            name, (price, _) = list(active_sources.items())[0]
            logger.info(f"یک منبع فعال برای {symbol}: {name} → {Utils.fmt_num(price)}")
            return price, 0.25, active_sources, None

        total_w = sum(w for _, w in active_sources.values())
        weighted_price = sum(p * (w / total_w) for p, w in active_sources.values())
        reliability = len(active_sources) / len(tasks)
        
        arz_price = None
        if 'arzdigital' in active_sources:
            arz_price = active_sources['arzdigital'][0]
        
        return float(weighted_price), float(reliability), active_sources, arz_price
        
    except Exception as e:
        logger.error(f"خطا در وزن‌دهی قیمت برای {symbol}: {e}")
        return None, 0.0, {}, None

# ========== ENTRY POINT ADJUSTMENT WITH ARZDIGITAL ==========
def calculate_entry_point_with_arz_premium(current_price: float, arz_price: float = None, 
                                         symbol: str = "") -> Tuple[float, float]:
    """محاسبه نقطه ورود با درنظرگیری تفاوت قیمت ArzDigital"""
    if arz_price is None or arz_price <= 0:
        return current_price, 0.0
    
    price_diff_percent = ((arz_price - current_price) / current_price) * 100
    
    symbol_adjustments = {
        "BTC/USDT": {"max_diff": 5.0, "adjustment": 0.4},
        "ETH/USDT": {"max_diff": 6.0, "adjustment": 0.5},
        "BNB/USDT": {"max_diff": 8.0, "adjustment": 0.6},
        "SOL/USDT": {"max_diff": 10.0, "adjustment": 0.7},
        "ADA/USDT": {"max_diff": 12.0, "adjustment": 0.7},
        "XRP/USDT": {"max_diff": 15.0, "adjustment": 0.8},
        "DOT/USDT": {"max_diff": 12.0, "adjustment": 0.7}
       
    }
    
    cfg = symbol_adjustments.get(symbol, {"max_diff": 10.0, "adjustment": 0.6})
    
    if abs(price_diff_percent) <= cfg["max_diff"] and price_diff_percent > 1.0:
        adjusted_price = current_price + ((arz_price - current_price) * cfg["adjustment"])
        
        logger.info(f"🔁 تفاوت قیمت {symbol}: {price_diff_percent:.2f}% | "
                   f"تعدیل نقطه ورود: {current_price:.2f} → {adjusted_price:.2f} "
                   f"(عامل تعدیل: {cfg['adjustment']})")
        return adjusted_price, price_diff_percent
    
    elif price_diff_percent > cfg["max_diff"]:
        logger.warning(f"⚠️  تفاوت قیمت {symbol} بسیار زیاد است: {price_diff_percent:.2f}% "
                      f"(حداکثر مجاز: {cfg['max_diff']}%). استفاده از قیمت پایه.")
    
    return current_price, price_diff_percent

# ========== NEWS FETCHERS ==========
class NewsFetcher:
    """دریافت و تحلیل اخبار"""
    
    def __init__(self, config):
        self.config = config
    
    @staticmethod
    def simple_sentiment(text: str) -> float:
        """تحلیل ساده احساسات"""
        txt = (text or "").lower()
        pos_words = ['rise', 'bull', 'gain', 'positive', 'up', 'surge', 'pump', 'rally']
        neg_words = ['fall', 'bear', 'loss', 'negative', 'down', 'dump', 'plunge']
        
        score = 0.0
        score += sum(1 for w in pos_words if w in txt) * 0.2
        score -= sum(1 for w in neg_words if w in txt) * 0.2
        
        return max(-1.0, min(1.0, score))
    
    @staticmethod
    def recency_boost(published_at: Optional[str]) -> float:
        """افزایش امتیاز براساس تازگی"""
        try:
            if not published_at:
                return 0.0
            
            dt = pd.to_datetime(published_at, utc=True)
            hours = (pd.Timestamp.utcnow() - dt).total_seconds() / 3600
            
            if hours <= 6:
                return 0.3
            elif hours <= 24:
                return 0.2
            elif hours <= 72:
                return 0.1
            
            return 0.0
        except Exception:
            return 0.0
    
    async def fetch_newsapi(self, session: aiohttp.ClientSession, symbol: str) -> Tuple[List[Dict], int]:
        """دریافت اخبار از NewsAPI"""
        try:
            url = "https://newsapi.org/v2/everything"
            params = {
                "q": f"{symbol.split('/')[0]} cryptocurrency",
                "apiKey": self.config.NEWSAPI_KEY,
                "pageSize": 10,
                "sortBy": "publishedAt",
                "language": "en",
            }
            
            data = await http_get(session, url, params=params, timeout=10)
            if data:
                articles = data.get("articles", [])
                scored_articles = []
                
                for art in articles:
                    title_lower = art.get('title', '').lower()
                    description_lower = art.get('description', '').lower()
                    content = title_lower + ' ' + description_lower
                    
                    score = 0.0
                    if any(w in content for w in ['crypto', 'bitcoin', 'ethereum']):
                        score += 0.5
                    
                    if any(word in content for word in ['rise', 'bullish', 'gain', 'positive', 'up']):
                        score += 0.2
                    elif any(word in content for word in ['fall', 'bearish', 'loss', 'negative', 'down']):
                        score -= 0.2
                    
                    item = {'title': art.get('title', ''), 'score': score}
                    if art.get('publishedAt'):
                        item['publishedAt'] = art['publishedAt']
                    
                    scored_articles.append(item)
                
                return scored_articles, len(articles)
                
        except Exception as e:
            logger.debug(f"NewsAPI failed for {symbol}: {e}")
        
        return [], 0
    
    async def fetch_cryptopanic(self, session: aiohttp.ClientSession, symbol: str) -> Tuple[List[Dict], int]:
        """دریافت اخبار از CryptoPanic"""
        try:
            url = "https://cryptopanic.com/api/v1/posts/"
            params = {
                "auth_token": self.config.CRYPTOPANIC_API_KEY,
                "currencies": symbol.split("/")[0],
                "kind": "news",
            }
            
            data = await http_get(session, url, params=params, timeout=10)
            if data:
                articles = data.get("results", [])
                scored_articles = []
                
                for art in articles:
                    title_lower = (art.get('title') or '').lower()
                    score = 0.0
                    
                    if 'crypto' in title_lower:
                        score += 0.5
                    
                    votes = art.get('votes', {})
                    if votes.get('positive', 0) > votes.get('negative', 0):
                        score += 0.3
                    
                    item = {'title': art.get('title', ''), 'score': score}
                    if art.get('published_at'):
                        item['published_at'] = art['published_at']
                    
                    scored_articles.append(item)
                
                return scored_articles, len(articles)
                
        except Exception as e:
            logger.debug(f"CryptoPanic failed for {symbol}: {e}")
        
        return [], 0
    
    async def fetch_coingecko_news(self, session: aiohttp.ClientSession, symbol: str) -> Tuple[List[Dict], int]:
        """دریافت اخبار از CoinGecko"""
        try:
            coin_map = {
                'BTC/USDT': 'bitcoin', 'ETH/USDT': 'ethereum', 'BNB/USDT': 'binancecoin',
                'ADA/USDT': 'cardano', 'SOL/USDT': 'solana', 'XRP/USDT': 'ripple',
                'DOT/USDT': 'polkadot'
            }
            
            coin_id = coin_map.get(symbol)
            if not coin_id:
                return [], 0
            
            url = f"https://api.coingecko.com/api/v3/coins/{coin_id}"
            params = {
                "tickers": "false",
                "market_data": "false",
                "community_data": "false",
                "developer_data": "false",
                "sparkline": "false",
            }
            
            headers = {"x-cg-demo-api-key": self.config.COINGECKO_API_KEY} if self.config.COINGECKO_API_KEY else {}
            data = await http_get(session, url, params=params, headers=headers, timeout=10)
            
            if data:
                links = data.get("links", {})
                articles = (links.get("homepage", []) or []) + (links.get("announcement_url", []) or [])
                scored_articles = []
                
                for link in articles:
                    s = 0.1
                    ll = (link or '').lower()
                    if 'crypto' in ll:
                        s += 0.4
                    scored_articles.append({'title': link, 'score': s})
                
                return scored_articles, len(articles)
                
        except Exception as e:
            logger.debug(f"CoinGecko news failed for {symbol}: {e}")
        
        return [], 0
    
    async def fetch_total_news(self, session: aiohttp.ClientSession, symbol: str) -> Tuple[int, float, float]:
        """دریافت و ترکیب اخبار از همه منابع"""
        newsapi_articles, newsapi_count = await self.fetch_newsapi(session, symbol)
        cryptopanic_articles, cryptopanic_count = await self.fetch_cryptopanic(session, symbol)
        coingecko_articles, coingecko_count = await self.fetch_coingecko_news(session, symbol)
        
        def normalize_items(items, source_name):
            normalized = []
            for it in items:
                title = it.get('title', '')
                score = it.get('score', 0.0)
                sentiment = self.simple_sentiment(title)
                published_at = it.get('publishedAt') or it.get('published_at') or None
                rec_boost = self.recency_boost(published_at)
                total_item_score = max(-1.0, min(1.0, score + sentiment + rec_boost))
                normalized.append({'source': source_name, 'title': title, 'score': total_item_score})
            return normalized
        
        all_items = []
        all_items += normalize_items(newsapi_articles, 'newsapi')
        all_items += normalize_items(cryptopanic_articles, 'cryptopanic')
        all_items += normalize_items(coingecko_articles, 'coingecko')
        
        total = 0.0
        total_weight = 0.0
        source_success = 0
        
        for src in ['newsapi', 'cryptopanic', 'coingecko']:
            src_items = [i for i in all_items if i['source'] == src]
            if src_items:
                source_success += 1
                src_avg = np.mean([i['score'] for i in src_items]) if src_items else 0.0
                w = self.config.NEWS_SOURCE_WEIGHTS.get(src, 0.2)
                total += src_avg * w
                total_weight += w
        
        news_score = (total / total_weight) if total_weight > 0 else 0.0
        total_news = newsapi_count + cryptopanic_count + coingecko_count
        news_reliability = source_success / 3.0
        
        logger.info(
            f"📰 اخبار {symbol} → تعداد={total_news} | "
            f"اتکا={news_reliability*100:.1f}% | امتیاز={news_score:.3f}"
        )
        
        return total_news, news_reliability, news_score

news_fetcher = NewsFetcher(config)

# ========== TECHNICAL ANALYZER ==========
class AdvancedTechnicalAnalyzer:
    """تحلیلگر تکنیکال پیشرفته"""
    
    def __init__(self, df: pd.DataFrame = None):
        self.df = df
        self.base_weights = {
            'macd': 0.15, 'rsi': 0.10, 'fibonacci': 0.10, 'volume': 0.10,
            'atr': 0.10, 'candlestick': 0.10, 'ichimoku': 0.10, 'divergence': 0.10,
            'adx': 0.05, 'bollinger': 0.05, 'ema_cross': 0.05, 'news': 0.10,
            'harmonic': 0.05, 'obv': 0.04, 'vwap': 0.04, 'supertrend': 0.05, 'psar': 0.04
        }
        self.cache = TTLCache(maxsize=200, ttl=300)

    def set_data(self, df: pd.DataFrame):
        self.df = df

    def adjust_weights_dynamically(self, price_reliability: float, news_reliability: float) -> Dict[str, float]:
        """تنظیم وزن‌ها به صورت پویا"""
        weights = self.base_weights.copy()
        
        if price_reliability > 0.7:
            weights['macd'] += 0.03
            weights['rsi'] += 0.03
            weights['ema_cross'] += 0.02
        
        if news_reliability > 0.7:
            weights['news'] += 0.05
        elif news_reliability < 0.3:
            weights['news'] = 0.0
        
        total = sum(weights.values())
        return {k: v / total for k, v in weights.items()}

    def detect_divergence(self, indicator: str = 'RSI') -> float:
        """تشخیص واگرایی"""
        close = self.df['close']
        
        if indicator.upper() == 'RSI':
            ind = talib.RSI(close, timeperiod=14)
        elif indicator.upper() == 'MACD':
            macd, macdsignal, _ = talib.MACD(close)
            ind = macd - macdsignal
        else:
            return 0.0
        
        score = 0.0
        for i in range(len(close) - 6, len(close) - 1):
            price1, price2 = close.iloc[i], close.iloc[i + 1]
            ind1, ind2 = ind.iloc[i], ind.iloc[i + 1]
            
            if price2 > price1 and ind2 < ind1:
                score -= 0.3
            elif price2 < price1 and ind2 > ind1:
                score += 0.3
            
            if price2 < price1 and ind2 < ind1:
                score += 0.2
            elif price2 > price1 and ind2 > ind1:
                score -= 0.2
        
        return score

    def calculate_macd_signal(self) -> float:
        """محاسبه سیگنال MACD"""
        try:
            close = self.df['close']
            macd, macd_signal, macd_hist = talib.MACD(close, fastperiod=12, slowperiod=26, signalperiod=9)
            current_hist = Utils.safe_get(macd_hist, -1, 0)
            prev_hist = Utils.safe_get(macd_hist, -2, 0)
            
            if current_hist > 0 and current_hist > prev_hist:
                return 1.0
            elif current_hist < 0 and current_hist < prev_hist:
                return -1.0
            elif current_hist > 0:
                return 0.5
            elif current_hist < 0:
                return -0.5
            return 0.0
        except Exception:
            return 0.0

    def calculate_rsi_signal(self) -> float:
        """محاسبه سیگنال RSI"""
        try:
            close = self.df['close']
            rsi = talib.RSI(close, timeperiod=14)
            current_rsi = Utils.safe_get(rsi, -1, 50)
            
            if current_rsi < 30:
                return 1.0
            elif current_rsi > 70:
                return -1.0
            elif current_rsi < 40:
                return 0.5
            elif current_rsi > 60:
                return -0.5
            return 0.0
        except Exception:
            return 0.0

    def calculate_volume_signal(self) -> float:
        """محاسبه سیگنال حجم"""
        try:
            volume = self.df['volume']
            current_volume = Utils.safe_get(volume, -1, 0)
            avg_volume = volume.tail(20).mean() if len(volume) >= 20 else current_volume
            volume_ratio = current_volume / avg_volume if avg_volume > 0 else 1
            
            if volume_ratio > 2.0:
                return 1.0
            elif volume_ratio > 1.5:
                return 0.5
            elif volume_ratio < 0.5:
                return -0.5
            return 0.0
        except Exception:
            return 0.0

    def calculate_candlestick_patterns(self) -> float:
        """تشخیص الگوهای کندلی"""
        try:
            open_, high_, low_, close_ = self.df['open'], self.df['high'], self.df['low'], self.df['close']
            
            bullish_patterns = [
                talib.CDLHAMMER(open_, high_, low_, close_),
                talib.CDLENGULFING(open_, high_, low_, close_),
                talib.CDLMORNINGSTAR(open_, high_, low_, close_),
                talib.CDLPIERCING(open_, high_, low_, close_)
            ]
            
            bearish_patterns = [
                talib.CDLSHOOTINGSTAR(open_, high_, low_, close_),
                talib.CDLDARKCLOUDCOVER(open_, high_, low_, close_),
                talib.CDLEVENINGSTAR(open_, high_, low_, close_),
                talib.CDLHANGINGMAN(open_, high_, low_, close_)
            ]
            
            bullish_score = sum(1 for pattern in bullish_patterns if pattern.iloc[-1] > 0)
            bearish_score = sum(1 for pattern in bearish_patterns if pattern.iloc[-1] > 0)
            
            return (bullish_score - bearish_score) / 4.0
        except Exception:
            return 0.0

    def calculate_bollinger_signal(self) -> float:
        """محاسبه سیگنال بولینگر باند"""
        try:
            close = self.df['close']
            upper, middle, lower = talib.BBANDS(close, timeperiod=20, nbdevup=2, nbdevdn=2, matype=0)
            
            current_close = Utils.safe_get(close, -1, 0)
            current_upper = Utils.safe_get(upper, -1, current_close)
            current_lower = Utils.safe_get(lower, -1, current_close)
            
            if current_close > current_upper:
                return -1.0
            elif current_close < current_lower:
                return 1.0
            return 0.0
        except Exception:
            return 0.0

    def calculate_ema_cross_signal(self) -> float:
        """محاسبه سیگنال کراس EMA"""
        try:
            close = self.df['close']
            ema_short = talib.EMA(close, timeperiod=12)
            ema_long = talib.EMA(close, timeperiod=26)
            
            ema_short_current = Utils.safe_get(ema_short, -1, 0)
            ema_short_prev = Utils.safe_get(ema_short, -2, 0)
            ema_long_current = Utils.safe_get(ema_long, -1, 0)
            ema_long_prev = Utils.safe_get(ema_long, -2, 0)
            
            if ema_short_current > ema_long_current and ema_short_prev <= ema_long_prev:
                return 1.0
            elif ema_short_current < ema_long_current and ema_short_prev >= ema_long_prev:
                return -1.0
            return 0.0
        except Exception:
            return 0.0

    def detect_harmonic_pattern(self) -> List[str]:
        """تشخیص الگوهای هارمونیک"""
        patterns = []
        try:
            close = self.df['close']
            high = self.df['high']
            low = self.df['low']
            
            # تشخیص ساده الگوها بر اساس حرکت قیمت
            recent_change = (close.iloc[-1] - close.iloc[-5]) / close.iloc[-5] * 100
            
            if abs(recent_change) > 10:
                if recent_change > 0:
                    patterns.append("Potential Bullish Pattern")
                else:
                    patterns.append("Potential Bearish Pattern")
            
        except Exception:
            pass
        
        return patterns

    def harmonic_validity_filter(self, patterns: List[str]) -> float:
        """فیلتر اعتبارسنجی الگوهای هارمونیک"""
        if not patterns:
            return 0.0
        
        valid_patterns = ['Potential Bullish Pattern', 'Potential Bearish Pattern']
        score = 0.0
        
        for pattern in patterns:
            if pattern in valid_patterns:
                if "Bullish" in pattern:
                    score += 0.7
                elif "Bearish" in pattern:
                    score -= 0.7
        
        return score

    def calculate_sl_tp(self, signal: str) -> Tuple[Optional[float], Optional[float]]:
        """محاسبه استاپ لاس و تیک پروفیت"""
        try:
            atr = Utils.calculate_atr(self.df)
            entry_price = float(self.df['close'].iloc[-1])
            
            # Swing High/Low
            swing_high = self.df['high'].rolling(20).max().iloc[-1]
            swing_low = self.df['low'].rolling(20).min().iloc[-1]
            
            if signal in ['BUY', 'STRONG_BUY']:
                stop_loss = max(swing_low - 0.5 * atr, entry_price * 0.98)
                take_profit = entry_price + (entry_price - stop_loss) * 1.5
                return float(stop_loss), float(take_profit)
            
            elif signal in ['SELL', 'STRONG_SELL']:
                stop_loss = min(swing_high + 0.5 * atr, entry_price * 1.02)
                take_profit = entry_price - (stop_loss - entry_price) * 1.5
                return float(stop_loss), float(take_profit)
            
            return None, None
            
        except Exception as e:
            logger.error(f"خطا در محاسبه SL/TP: {e}")
            return None, None

    def comprehensive_analysis(self, price_reliability: float = 1.0, 
                             news_reliability: float = 1.0, 
                             news_score: float = 0.0) -> Dict[str, Any]:
        """تحلیل جامع تکنیکال"""
        if self.df is None or self.df.empty:
            return {"signal": "HOLD", "score": 0.0, "confidence": 0, "indicators": {}, 
                    "summary": {"overall_signal": "HOLD", "confidence": 0}}
        
        try:
            # محاسبه اندیکاتورها
            indicators = {
                'macd': self.calculate_macd_signal(),
                'rsi': self.calculate_rsi_signal(),
                'volume': self.calculate_volume_signal(),
                'candlestick': self.calculate_candlestick_patterns(),
                'divergence': self.detect_divergence('RSI'),
                'bollinger': self.calculate_bollinger_signal(),
                'ema_cross': self.calculate_ema_cross_signal(),
                'news': news_score,
            }
            
            # تشخیص الگوهای هارمونیک
            patterns = self.detect_harmonic_pattern()
            indicators['harmonic'] = self.harmonic_validity_filter(patterns)
            
            # فیبوناچی
            try:
                high = self.df['high'].max()
                low = self.df['low'].min()
                close = self.df['close'].iloc[-1]
                fib_score = 0.5 if close > (high + low) / 2 else -0.5
                indicators['fibonacci'] = fib_score
            except Exception as e:
                logger.warning(f"محاسبه فیبوناچی ناموفق: {e}")
                indicators['fibonacci'] = 0.0
            
            # وزندهی پویا
            weights = self.adjust_weights_dynamically(price_reliability, news_reliability)
            base_score = sum(indicators[k] * weights.get(k, 0.0) for k in indicators)
            base_score = max(-1.0, min(1.0, base_score))
            
            reliability_factor = (price_reliability + news_reliability) / 2
            final_score = base_score * reliability_factor
            
            # تولید سیگنال نهایی
            if final_score > 0.3:
                signal, confidence = "STRONG_BUY", min(100, final_score * 150)
            elif final_score > 0.1:
                signal, confidence = "BUY", min(80, final_score * 120)
            elif final_score < -0.3:
                signal, confidence = "STRONG_SELL", min(100, abs(final_score) * 150)
            elif final_score < -0.1:
                signal, confidence = "SELL", min(80, abs(final_score) * 120)
            else:
                signal, confidence = "HOLD", 0
            
            # محاسبه SL/TP
            stop_loss, take_profit = self.calculate_sl_tp(signal)
            
            return {
                'signal': signal,
                'score': round(final_score, 3),
                'confidence': round(confidence, 2),
                'indicators': indicators,
                'harmonic_patterns': patterns,
                'reliability': {
                    'price': round(price_reliability, 3),
                    'news': round(news_reliability, 3),
                    'overall': round(reliability_factor, 3)
                },
                'stop_loss': stop_loss,
                'take_profit': take_profit
            }
            
        except Exception as e:
            logger.error(f"خطا در تحلیل تکنیکال: {e}")
            return {
                'signal': 'HOLD',
                'score': 0.0,
                'confidence': 0,
                'indicators': {},
                'harmonic_patterns': [],
                'reliability': {},
                'stop_loss': None,
                'take_profit': None
            }

# ========== RULE ENGINE ==========
class UnifiedStrategy:
    """کلاس سیگنال Rule-based"""
    
    def __init__(self, symbol: str, timeframe: str, side: str, entry: float, 
                 sl: float, tp: float, confidence: float, rule: str):
        self.symbol = symbol
        self.timeframe = timeframe
        self.side = side  # "BUY" یا "SELL"
        self.entry = entry
        self.sl = sl
        self.tp = tp
        self.confidence = confidence
        self.rule = rule

def generate_rule_signals(df_rule: pd.DataFrame, symbol: str, timeframe: str = "") -> List[UnifiedStrategy]:
    """تولید سیگنال‌های مبتنی بر قواعد"""
    signals: List[UnifiedStrategy] = []
    try:
        close = df_rule["Close"]
        ema_short = talib.EMA(close, timeperiod=12)
        ema_long = talib.EMA(close, timeperiod=26)
        rsi = talib.RSI(close, timeperiod=14)

        bullish_cross = ema_short.iloc[-1] > ema_long.iloc[-1] and ema_short.iloc[-2] <= ema_long.iloc[-2]
        bearish_cross = ema_short.iloc[-1] < ema_long.iloc[-1] and ema_short.iloc[-2] >= ema_long.iloc[-2]

        rsi_bull_ok = rsi.iloc[-1] < 65
        rsi_bear_ok = rsi.iloc[-1] > 35

        entry = float(close.iloc[-1])
        swing_high = float(df_rule["High"].rolling(20).max().iloc[-1])
        swing_low = float(df_rule["Low"].rolling(20).min().iloc[-1])
        atr_proxy = float((df_rule["High"] - df_rule["Low"]).tail(14).mean())

        if bullish_cross and rsi_bull_ok:
            sl = max(swing_low - 0.5 * atr_proxy, entry * 0.98)
            tp = entry + (entry - sl) * 1.5
            signals.append(UnifiedStrategy(symbol, timeframe or "", "BUY", entry, sl, tp, 70.0, "EMA12/26 + RSI"))

        elif bearish_cross and rsi_bear_ok:
            sl = min(swing_high + 0.5 * atr_proxy, entry * 1.02)
            tp = entry - (sl - entry) * 1.5
            signals.append(UnifiedStrategy(symbol, timeframe or "", "SELL", entry, sl, tp, 70.0, "EMA12/26 + RSI"))

        else:
            if rsi.iloc[-1] < 30:
                sl = max(swing_low - 0.5 * atr_proxy, entry * 0.98)
                tp = entry + (entry - sl) * 1.2
                signals.append(UnifiedStrategy(symbol, timeframe or "", "BUY", entry, sl, tp, 55.0, "RSI<30"))
            elif rsi.iloc[-1] > 70:
                sl = min(swing_high + 0.5 * atr_proxy, entry * 1.02)
                tp = entry - (sl - entry) * 1.2
                signals.append(UnifiedStrategy(symbol, timeframe or "", "SELL", entry, sl, tp, 55.0, "RSI>70"))

    except Exception as e:
        logger.debug(f"خطای Rule signals برای {symbol}: {e}")

    return signals

def to_rule_df(df_coingecko: pd.DataFrame) -> pd.DataFrame:
    """تبدیل DataFrame به فرمت Rule-compatible"""
    df = df_coingecko.copy()
    df = df.rename(columns={
        'open': 'Open', 
        'high': 'High', 
        'low': 'Low', 
        'close': 'Close', 
        'volume': 'Volume'
    })
    df['Date'] = df.index
    df = df.reset_index(drop=True)
    df = df.sort_values('Date').reset_index(drop=True)
    
    for col in ['Open', 'High', 'Low']:
        if col not in df.columns:
            df[col] = df['Close']
    if 'Volume' not in df.columns:
        df['Volume'] = np.nan
        
    return df[['Date', 'Open', 'High', 'Low', 'Close', 'Volume']]

# ========== DATA FETCHER ==========
class DataFetcher:
    """دریافت داده‌های OHLCV"""
    
    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.cache = TTLCache(maxsize=100, ttl=300)
    
    async def fetch_ohlcv(self, symbol: str, timeframe: str = "1h", limit: int = 150) -> pd.DataFrame:
        """دریافت داده‌های OHLCV از CoinGecko"""
        cache_key = f"{symbol}_{timeframe}_{limit}"
        if cache_key in self.cache:
            return self.cache[cache_key]
        
        try:
            coin_map = {
                'BTC/USDT': 'bitcoin', 'ETH/USDT': 'ethereum', 'BNB/USDT': 'binancecoin',
                'ADA/USDT': 'cardano', 'SOL/USDT': 'solana', 'XRP/USDT': 'ripple',
                'DOT/USDT': 'polkadot'

            }
            
            coin_id = coin_map.get(symbol)
            if not coin_id:
                return pd.DataFrame()
            
            tf_days = {'15m': 7, '30m': 14, '1h': 30, '4h': 60, '1d': 90}
            days = tf_days.get(timeframe, 30)
            
            url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart"
            params = {'vs_currency': 'usd', 'days': days}
            headers = {'x-cg-demo-api-key': config.COINGECKO_API_KEY} if config.COINGECKO_API_KEY else {}
            
            data = await http_get(self.session, url, params=params, headers=headers, timeout=20)
            if not data or 'prices' not in data or 'total_volumes' not in data:
                return pd.DataFrame()
            
            # پردازش داده‌های قیمت
            prices = pd.DataFrame(data['prices'], columns=['timestamp', 'price'])
            prices['timestamp'] = pd.to_datetime(prices['timestamp'], unit='ms', utc=True)
            prices = prices.set_index('timestamp')
            
            # تبدیل به OHLCV
            freq_map = {'15m': '15min', '30m': '30min', '1h': '1h', '4h': '4h', '1d': '1D'}
            freq = freq_map.get(timeframe, '1h')
            ohlc = prices['price'].resample(freq).ohlc()
            
            # پردازش حجم
            volumes = pd.DataFrame(data['total_volumes'], columns=['timestamp', 'volume'])
            volumes['timestamp'] = pd.to_datetime(volumes['timestamp'], unit='ms', utc=True)
            volumes = volumes.set_index('timestamp')
            ohlc['volume'] = volumes['volume'].resample(freq).sum()
            
            df = ohlc.dropna().tail(limit)
            self.cache[cache_key] = df
            
            return df
            
        except Exception as e:
            logger.error(f"Error fetching OHLCV for {symbol}: {e}")
        
        return pd.DataFrame()

# ========== RISK MANAGER ==========
class RiskManager:
    """مدیریت ریسک"""
    
    def __init__(self, config):
        self.config = config
    
    def calculate_position_size(self, signal: Dict, account_balance: float) -> float:
        """محاسبه اندازه پوزیشن"""
        confidence = signal.get('confidence', 0)
        base_risk = self.config.RISK_PER_TRADE
        
        # تنظیم ضریب ریسک براساس اطمینان
        risk_multiplier = 1.0
        if confidence >= self.config.STRONG_SIGNAL_THRESHOLD:
            risk_multiplier = 1.2
        elif confidence >= 70:
            risk_multiplier = 1.0
        elif confidence >= self.config.MIN_SIGNAL_CONFIDENCE:
            risk_multiplier = 0.7
        
        final_risk = base_risk * risk_multiplier
        final_risk = min(final_risk, self.config.MAX_POSITION_SIZE)
        
        # محاسبه براساس استاپ لاس
        stop_loss = signal.get('stop_loss', 0)
        current_price = signal.get('current_price', 0)
        
        if stop_loss and current_price:
            risk_per_unit = abs(current_price - stop_loss)
            if risk_per_unit > 0:
                units = (account_balance * final_risk) / risk_per_unit
                return min(units, (account_balance * final_risk) / current_price)
        
        return (account_balance * final_risk) / current_price if current_price > 0 else 0

    def validate_risk_parameters(self, signal):
        """اعتبارسنجی پارامترهای ریسک"""
        sl = signal.get('stop_loss', 0)
        tp = signal.get('take_profit', 0)
        price = signal.get('current_price', 0)
        
        if not all([sl, tp, price]):
            return False, "پارامترهای ریسک ناقص"
        
        risk = abs(price - sl)
        reward = abs(tp - price)
        risk_reward_ratio = reward / risk if risk > 0 else 0
        
        if risk_reward_ratio < 1.2:
            return False, f"نسبت Risk/Reward نامناسب: {risk_reward_ratio:.2f}"
        
        stop_loss_percent = abs(price - sl) / price * 100
        if stop_loss_percent > 10:
            return False, f"استاپ لاس بسیار بزرگ: {stop_loss_percent:.1f}%"
        
        return True, "پارامترهای ریسک معتبر"

risk_manager = RiskManager(config)

# ========== SIGNAL FILTER ==========
class SignalFilter:
    """فیلترهای سیگنال"""
    
    def __init__(self, config):
        self.config = config
    
    def apply_filters(self, signals: List[Dict], account_balance: float = 1000) -> List[Dict]:
        """اعمال فیلترهای نهایی"""
        filtered_signals = []
        
        for signal in signals:
            # فیلتر 1: حداقل اطمینان
            if signal.get('confidence', 0) < self.config.MIN_SIGNAL_CONFIDENCE:
                continue
            
            # فیلتر 2: اعتبارسنجی ریسک
            if signal.get('stop_loss') and signal.get('take_profit'):
                is_risk_valid, risk_message = risk_manager.validate_risk_parameters(signal)
                if not is_risk_valid:
                    logger.warning(f"❌ رد سیگنال به دلیل ریسک: {signal.get('symbol', '')} - {risk_message}")
                    continue
            
            # فیلتر 3: محاسبه اندازه پوزیشن
            position_size = risk_manager.calculate_position_size(signal, account_balance)
            if position_size <= 0:
                continue
            
            # افزودن اطلاعات تکمیلی
            signal['position_size'] = position_size
            signal['risk_percentage'] = self.config.RISK_PER_TRADE * 100
            
            filtered_signals.append(signal)
        
        # مرتب‌سازی بر اساس اعتماد
        filtered_signals.sort(key=lambda x: x.get('confidence', 0), reverse=True)
        
        logger.info(f"📊 قبل از فیلتر: {len(signals)} سیگنال")
        logger.info(f"📊 بعد از فیلتر: {len(filtered_signals)} سیگنال")
        
        if filtered_signals:
            avg_confidence = sum(s.get('confidence', 0) for s in filtered_signals) / len(filtered_signals)
            logger.info(f"📈 میانگین اطمینان: {avg_confidence:.1f}%")
            
            # ثبت متریک
            if SIGNAL_QUALITY:
                SIGNAL_QUALITY.set(avg_confidence)
        
        return filtered_signals

signal_filter = SignalFilter(config)

# ========== NOTIFICATION MANAGER ==========
class NotificationManager:
    """مدیریت نوتیفیکیشن‌ها"""
    
    def __init__(self, config):
        self.config = config
        self.email_manager = EmailManager(config)
        self.sms_manager = SMSManager(config)
    
    async def send_all_notifications(self, signal: Dict) -> Dict[str, bool]:
        """ارسال همه نوتیفیکیشن‌ها"""
        results = {
            'email': False,
            'sms': False
        }
        
        # بررسی آستانه برای ارسال
        confidence = signal.get('confidence', 0)
        
        # ارسال ایمیل
        if (self.config.FEATURE_FLAGS.get('email_alerts') and 
            confidence >= self.config.SMS_THRESHOLD):
            subject, body, html_body = self.email_manager.format_signal_email(signal)
            results['email'] = self.email_manager.send_email(subject, body, html_body)
        
        # ارسال SMS
        if (self.config.FEATURE_FLAGS.get('sms_alerts') and 
            confidence >= self.config.SMS_THRESHOLD):
            sms_message = self.sms_manager.format_signal_sms(signal)
            sms_result = self.sms_manager.send_sms(sms_message)
            results['sms'] = len(sms_result['success']) > 0
        
        # لاگ نتایج
        self.log_notification_results(signal, results)
        
        return results
    
    def log_notification_results(self, signal: Dict, results: Dict[str, bool]):
        """ثبت نتایج ارسال نوتیفیکیشن"""
        try:
            with get_db_connection() as conn:
                cur = conn.cursor()
                
                for platform, success in results.items():
                    if success:  # فقط در صورت موفقیت ثبت کن
                        cur.execute("""
                            INSERT INTO notification_logs 
                            (platform, symbol, confidence, message, success)
                            VALUES (?, ?, ?, ?, ?)
                        """, (
                            platform.upper(),
                            signal.get('symbol', ''),
                            signal.get('confidence', 0),
                            f"Signal {signal.get('signal', '')}",
                            success
                        ))
                        
        except Exception as e:
            logger.error(f"خطا در ثبت نتایج نوتیفیکیشن: {e}")

notification_manager = NotificationManager(config)

# ========== MAIN ANALYZER ==========
async def analyze_symbol(
    symbol: str,
    timeframe: str,
    session: aiohttp.ClientSession,
    news_cache: TTLCache,
    price_cache: Dict[str, Tuple[Optional[float], float, Optional[float]]]
) -> Optional[Dict]:
    """تحلیل کامل یک نماد"""
    logger.info(f"🔍 شروع تحلیل {symbol} روی تایم‌فریم {timeframe}")
    
    try:
        # دریافت قیمت
        if symbol in price_cache:
            current_price, price_reliability, arz_price = price_cache[symbol]
        else:
            current_price, price_reliability, _, arz_price = await fetch_price_weighted(session, symbol)
            price_cache[symbol] = (current_price, price_reliability, arz_price)

        if current_price is None:
            logger.warning(f"❌ دریافت قیمت ناموفق برای {symbol}")
            return None

        # تعدیل نقطه ورود با ArzDigital
        adjusted_entry = current_price
        price_diff_percent = 0.0
        
        if arz_price and arz_price > 0 and current_price and current_price > 0:
            adjusted_entry, price_diff_percent = calculate_entry_point_with_arz_premium(
                current_price, arz_price, symbol
            )

        # دریافت اخبار
        cache_key = f"news_{symbol}"
        if cache_key in news_cache:
            news_count, news_reliability, news_score = news_cache[cache_key]
        else:
            news_count, news_reliability, news_score = await news_fetcher.fetch_total_news(session, symbol)
            news_cache[cache_key] = (news_count, news_reliability, news_score)

        # دریافت داده OHLCV
        data_fetcher = DataFetcher(session)
        df_ohlcv = await data_fetcher.fetch_ohlcv(symbol, timeframe, 150)
        
        if df_ohlcv.empty or len(df_ohlcv) < 50:
            logger.warning(f"داده OHLCV ناکافی برای {symbol}")
            return None

        # تحلیل تکنیکال
        analyzer = AdvancedTechnicalAnalyzer(df_ohlcv)
        analysis = analyzer.comprehensive_analysis(price_reliability, news_reliability, news_score)

        # Rule-based signals
        df_rule = to_rule_df(df_ohlcv)
        rule_signals = generate_rule_signals(df_rule, symbol=symbol, timeframe=timeframe)

        # تلفیق نتایج
        entry_price = adjusted_entry
        rule_side = rule_entry = rule_sl = rule_tp = rule_conf = rule_rule = None

        if rule_signals:
            rs = rule_signals[0]
            rule_side = getattr(rs, 'side', None)
            rule_entry = getattr(rs, 'entry', None)
            
            if isinstance(rule_entry, (int, float)) and rule_entry > 0:
                combined_entry = (adjusted_entry + rule_entry) / 2
                if abs(combined_entry - adjusted_entry) / adjusted_entry < 0.05:
                    entry_price = combined_entry
                    logger.info(f"🔀 ترکیب نقطه ورود: Rule={rule_entry} + Arz-Adjusted={adjusted_entry} = {combined_entry}")
            
            rule_sl = getattr(rs, 'sl', None)
            rule_tp = getattr(rs, 'tp', None)
            rule_conf = float(getattr(rs, 'confidence', 0.0))
            rule_rule = getattr(rs, 'rule', None)

        # استفاده از SL/TP از تحلیل تکنیکال اگر rule-based نداشتیم
        final_sl = rule_sl if rule_sl is not None else analysis.get('stop_loss')
        final_tp = rule_tp if rule_tp is not None else analysis.get('take_profit')

        # سیگنال نهایی
        result = {
            'symbol': symbol,
            'timeframe': timeframe,
            'signal': analysis['signal'],
            'score': analysis['score'],
            'confidence': analysis['confidence'],
            'indicators': analysis['indicators'],
            'current_price': current_price,
            'entry_price': entry_price,
            'stop_loss': final_sl,
            'take_profit': final_tp,
            'news_count': news_count,
            'price_reliability': price_reliability,
            'news_reliability': news_reliability,
            'news_score': news_score,
            'timestamp': datetime.now().isoformat(),
            'rule_side': rule_side,
            'rule_entry': rule_entry,
            'rule_sl': rule_sl,
            'rule_tp': rule_tp,
            'rule_confidence': rule_conf,
            'rule_name': rule_rule,
            'ts': f"{symbol}-{timeframe}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            'arz_price': arz_price,
            'price_diff_percent': price_diff_percent,
            'entry_adjusted': adjusted_entry,
            'entry_original': current_price,
            'harmonic_patterns': analysis.get('harmonic_patterns', [])
        }

        if result['signal'] != 'HOLD' or rule_side:
            logger.info(
                f"📈 سیگنال قوی: {symbol} {timeframe} → {result['signal']} | "
                f"Conf={result['confidence']:.1f}% | "
                f"Price={current_price:.6f} | "
                f"Entry={entry_price:.6f} | "
                f"ArzDiff={price_diff_percent:.2f}% | "
                f"Rule: {rule_rule or '—'}"
            )

        return result

    except Exception as e:
        logger.error(f"❌ خطای تحلیل {symbol} {timeframe}: {e}", exc_info=True)
        return None

# ========== EXCEL REPORTER ==========
class ExcelReporter:
    """تولیدکننده گزارش Excel"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_report(self, signals: List[Dict[str, Any]]):
        if not signals:
            logger.warning("📭 هیچ سیگنالی برای گزارش اکسل")
            return
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.output_dir, f"crypto_signals_{timestamp}.xlsx")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "سیگنال‌ها"
            
            headers = [
                "نماد", "تایم‌فریم", "سیگنال", "اعتماد (%)", "امتیاز",
                "قیمت ورود", "حد ضرر", "حد سود", "قانون", "ورود قانون",
                "SL قانون", "TP قانون", "اعتماد قانون (%)", "قیمت ArzDigital",
                "تفاوت قیمت (%)", "ورود تعدیل شده", "تعداد خبر",
                "قابلیت اتکای قیمت (%)", "قابلیت اتکای خبر (%)", "تاریخ تحلیل"
            ]
            
            # استایل‌ها
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # هدرها
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # داده‌ها
            for row, s in enumerate(signals, start=2):
                sig = s.get("signal", "")

                # پاکسازی مقادیر عددی با Utils.safe_float
                conf = Utils.safe_float(s.get("confidence"), 0.0)
                score = Utils.safe_float(s.get("score"), 0.0)
                entry = Utils.safe_float(s.get("entry_price"), 0.0)
                sl = Utils.safe_float(s.get("stop_loss"), 0.0)
                tp = Utils.safe_float(s.get("take_profit"), 0.0)
                rule_conf = Utils.safe_float(s.get("rule_confidence"), 0.0)
                arz_price = Utils.safe_float(s.get("arz_price"), 0.0)
                arz_diff = Utils.safe_float(s.get("price_diff_percent"), 0.0)
                entry_adj = Utils.safe_float(s.get("entry_adjusted"), 0.0)
                price_rel = Utils.safe_float(s.get("price_reliability"), 0.0) * 100
                news_rel = Utils.safe_float(s.get("news_reliability"), 0.0) * 100

                ws.cell(row=row, column=1, value=s.get("symbol", ""))
                ws.cell(row=row, column=2, value=s.get("timeframe", ""))
                signal_cell = ws.cell(row=row, column=3, value=sig)
                ws.cell(row=row, column=4, value=round(conf, 2))
                ws.cell(row=row, column=5, value=round(score, 6))
                ws.cell(row=row, column=6, value=entry)
                ws.cell(row=row, column=7, value=sl)
                ws.cell(row=row, column=8, value=tp)
                ws.cell(row=row, column=9, value=s.get("rule_name", ""))
                ws.cell(row=row, column=10, value=Utils.safe_float(s.get("rule_entry"), 0.0))
                ws.cell(row=row, column=11, value=Utils.safe_float(s.get("rule_sl"), 0.0))
                ws.cell(row=row, column=12, value=Utils.safe_float(s.get("rule_tp"), 0.0))
                ws.cell(row=row, column=13, value=round(rule_conf, 2))
                ws.cell(row=row, column=14, value=arz_price)
                ws.cell(row=row, column=15, value=round(arz_diff, 2))
                ws.cell(row=row, column=16, value=entry_adj)
                ws.cell(row=row, column=17, value=s.get("news_count", 0))
                ws.cell(row=row, column=18, value=round(price_rel, 2))
                ws.cell(row=row, column=19, value=round(news_rel, 2))
                ws.cell(row=row, column=20, value=s.get("timestamp", ""))

                # رنگ‌آمیزی سیگنال
                if "BUY" in sig:
                    signal_cell.fill = buy_fill
                elif "SELL" in sig:
                    signal_cell.fill = sell_fill

            # تنظیم عرض ستون‌ها
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb.save(filename)
            logger.info(f"📄 فایل اکسل ذخیره شد: {filename}")
            
        except Exception as e:
            logger.error(f"❌ خطای ساخت اکسل: {e}")

# ========== HEALTH MONITOR ==========
class HealthMonitor:
    """مانیتور سلامت سیستم"""
    
    def __init__(self):
        self.start_time = datetime.now(timezone.utc)
    
    def get_health_status(self):
        return {
            "status": "healthy",
            "version": "6.0.0",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "uptime_seconds": (datetime.now(timezone.utc) - self.start_time).total_seconds(),
            "memory_usage_mb": self.get_memory_usage(),
            "active_symbols": len(config.SYMBOLS),
            "features": {
                "email_alerts": config.FEATURE_FLAGS.get('email_alerts', False),
                "sms_alerts": config.FEATURE_FLAGS.get('sms_alerts', False)
            }
        }
    
    def get_memory_usage(self):
        try:
            process = psutil.Process()
            return round(process.memory_info().rss / 1024 / 1024, 2)
        except Exception:
            return 0.0

# ========== MAIN ANALYSIS ==========
async def main_analysis() -> List[Dict[str, Any]]:
    """تابع اصلی تحلیل"""
    logger.info("🚀 شروع تحلیل جامع بازار رمزارز...")
    init_db()
    news_cache = TTLCache(maxsize=50, ttl=600)

    async with aiohttp.ClientSession(connector=TCPConnector(limit=20)) as session:
        # کش کردن قیمت‌ها
        price_cache = {}
        for sym in config.SYMBOLS:
            try:
                price, reliability, _, arz_price = await fetch_price_weighted(session, sym)
                price_cache[sym] = (price, reliability, arz_price)
                if price:
                    logger.info(f"💰 قیمت {sym}: {price:.6f} (اتکا: {reliability*100:.1f}%)")
                    if arz_price:
                        diff = ((arz_price - price) / price) * 100 if price > 0 else 0
                        logger.info(f"   ArzDigital: {arz_price:.6f} (تفاوت: {diff:.2f}%)")
            except Exception as e:
                logger.error(f"خطای قیمت {sym}: {e}")
                price_cache[sym] = (None, 0.0, None)

        # تحلیل همزمان
        semaphore = asyncio.Semaphore(10)
        
        async def analyze_with_limit(symbol, timeframe):
            async with semaphore:
                return await analyze_symbol(symbol, timeframe, session, news_cache, price_cache)

        tasks = [analyze_with_limit(sym, tf) for sym in config.SYMBOLS for tf in config.TIMEFRAMES]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # فیلتر سیگنال‌ها
        raw_signals = [
            r for r in results
            if isinstance(r, dict) and r.get("signal") and r.get("signal") != "HOLD"
        ]
        
        filtered_signals = signal_filter.apply_filters(raw_signals, config.INITIAL_BALANCE)
        filtered_signals.sort(key=lambda x: x.get("confidence", 0), reverse=True)
        top_signals = filtered_signals[:5]

        # ذخیره در دیتابیس
        for s in filtered_signals:
            try:
                with get_db_connection() as conn:
                    cur = conn.cursor()
                    cur.execute("""
                        INSERT OR IGNORE INTO signals (
                            ts, symbol, timeframe, signal, score, confidence,
                            price, sl, tp, news_score, price_rel, news_rel,
                            arz_price, price_diff_percent, entry_adjusted,
                            rule_name, rule_side, rule_entry, rule_sl, rule_tp, rule_confidence
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        s.get("ts"), s.get("symbol"), s.get("timeframe"), s.get("signal"),
                        s.get("score"), s.get("confidence"), s.get("entry_price"),
                        s.get("stop_loss"), s.get("take_profit"), s.get("news_score"),
                        s.get("price_reliability"), s.get("news_reliability"),
                        s.get("arz_price"), s.get("price_diff_percent"), s.get("entry_adjusted"),
                        s.get("rule_name"), s.get("rule_side"), s.get("rule_entry"),
                        s.get("rule_sl"), s.get("rule_tp"), s.get("rule_confidence")
                    ))
                logger.info(f"💾 ذخیره سیگنال: {s.get('symbol')} {s.get('timeframe')}")
            except Exception as e:
                logger.error(f"❌ خطای ذخیره سیگنال {s.get('symbol')}: {e}")

        # تولید گزارش Excel
        try:
            reporter = ExcelReporter(config.OUTPUT_DIR)
            reporter.generate_report(filtered_signals)
            logger.info("📊 گزارش اکسل تولید شد")
        except Exception as e:
            logger.error(f"❌ خطای گزارش اکسل: {e}")

        # ارسال نوتیفیکیشن‌ها برای سیگنال‌های برتر
        for signal in top_signals:
            if signal.get('confidence', 0) >= config.SMS_THRESHOLD:
                results = await notification_manager.send_all_notifications(signal)
                
                # لاگ نتایج
                platforms = []
                if results.get('email'):
                    platforms.append('Email')
                if results.get('sms'):
                    platforms.append('SMS')
                
                if platforms:
                    logger.info(f"📨 ارسال نوتیفیکیشن برای {signal.get('symbol')}: {', '.join(platforms)}")

        logger.info(f"✅ تحلیل کامل شد. {len(filtered_signals)} سیگنال معتبر")
        
        if top_signals:
            symbols_text = ", ".join(f"{s['symbol']} {s['timeframe']}" for s in top_signals)
            logger.info(f"🏆 سیگنال‌های برتر: {symbols_text}")

        # ثبت متریک
        if ACTIVE_SIGNALS:
            ACTIVE_SIGNALS.set(len(filtered_signals))
        
        if CACHE_HIT_RATE:
            CACHE_HIT_RATE.set(PRICE_CACHE.get_hit_rate())

        return top_signals

# ========== MAIN ENTRY ==========
async def run_periodically():
    """حالت اجرای دوره‌ای"""
    logger.info("🔄 حالت اجرای دوره‌ای آغاز شد...")
    init_db()
    
    while True:
        try:
            start = datetime.now(timezone.utc)
            logger.info(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] شروع تحلیل...")
            
            top = await main_analysis()
            
            elapsed = (datetime.now(timezone.utc) - start).total_seconds()
            sleep_for = max(1, config.RUN_INTERVAL - elapsed)
            
            logger.info(f"⏱️  زمان تحلیل: {elapsed:.1f} ثانیه | خواب برای: {sleep_for:.1f} ثانیه")
            await asyncio.sleep(sleep_for)
            
        except KeyboardInterrupt:
            logger.info("⏹️  توقف توسط کاربر")
            break
        except Exception as e:
            logger.error(f"❌ خطای حلقه اصلی: {e}", exc_info=True)
            await asyncio.sleep(60)

async def start_health_server():
    """شروع سرور سلامت"""
    app = web.Application()
    
    async def health_check(request):
        monitor = HealthMonitor()
        return web.json_response(monitor.get_health_status())
    
    async def metrics_handler(request):
        try:
            metrics_data = generate_latest()
            return web.Response(body=metrics_data, content_type="text/plain")
        except Exception as e:
            return web.json_response({"error": str(e)}, status=500)
    
    app.router.add_get("/health", health_check)
    app.router.add_get("/metrics", metrics_handler)
    
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", 8080)
    await site.start()
    
    logger.info("🌐 سرور سلامت روی http://0.0.0.0:8080 شروع شد")
    return runner

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="سیستم تحلیل رمزارز پیشرفته Leila Trading Bot Pro")
    parser.add_argument("--once", action="store_true", help="اجرای یکباره")
    parser.add_argument("--loop", action="store_true", help="اجرای دوره‌ای")
    parser.add_argument("--health", action="store_true", help="اجرای سرور سلامت")
    
    args = parser.parse_args()
    
    init_db()
    logger.info("🚀 سیستم تحلیل رمزارز Leila Trading Bot Pro آماده است")
    
    if args.health:
        asyncio.run(start_health_server())
    
    elif args.once:
        logger.info("▶️  اجرای یکباره تحلیل...")
        asyncio.run(main_analysis())
    
    elif args.loop:
        logger.info("🔄 شروع اجرای دوره‌ای...")
        asyncio.run(run_periodically())
    
    else:
        logger.info("▶️  اجرای پیش‌فرض (یکباره)...")
        asyncio.run(main_analysis())